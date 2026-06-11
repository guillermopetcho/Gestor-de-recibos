import os, time, subprocess
import platform
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.core import database

def generar_recibo_excel(recibo_id):
    recibo = database.obtener_recibo(recibo_id)
    if not recibo: raise ValueError("Recibo no encontrado")
    
    items = database.obtener_items_recibo(recibo_id)
    config = database.obtener_config()
    
    # Resolucion de rutas
    ruta_plantilla = recibo.get('ruta_plantilla', '')
    ruta_guardado = recibo.get('ruta_guardado', '')
    if not ruta_guardado: # Si el edificio no tiene, usar la global
        ruta_guardado = config.get("ruta_guardado", "recibos")
        
    import sys
    # Determinar la ruta base
    if getattr(sys, 'frozen', False):
        BASE_DIR = Path(sys.executable).parent
    else:
        BASE_DIR = Path(__file__).parent.parent.parent

    ruta_base = Path(ruta_guardado) if os.path.isabs(ruta_guardado) else BASE_DIR / ruta_guardado
    ruta_base.mkdir(parents=True, exist_ok=True)
    
    nombre_archivo = f"Recibo_{recibo['edificio_nombre'].replace(' ', '_')}_{recibo['identificador'].replace(' ', '_')}_{recibo['periodo'].replace('/', '-')}.xlsx"
    ruta_archivo = ruta_base / nombre_archivo

    # --- FLUJO: CON PLANTILLA PERSONALIZADA ---
    if ruta_plantilla and os.path.exists(ruta_plantilla):
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active
        
        reemplazos = {
            "{{INQUILINO}}": recibo.get("inquilino", "") or "",
            "{{DEPARTAMENTO}}": recibo.get("identificador", "") or "",
            "{{EDIFICIO}}": recibo.get("edificio_nombre", "") or "",
            "{{PERIODO}}": recibo.get("periodo", "") or "",
            "{{FECHA}}": str(recibo.get("fecha_emision", "")).split(" ")[0],
            "{{TOTAL}}": str(recibo.get("total", 0.0)),
        }
        
        mapeo_celdas = recibo.get("mapeo_celdas", "")
        if mapeo_celdas:
            import json
            try:
                mapeo = json.loads(mapeo_celdas)
            except:
                mapeo = {}
                
            if mapeo.get("inquilino"): ws[mapeo["inquilino"]] = reemplazos["{{INQUILINO}}"]
            if mapeo.get("departamento"): ws[mapeo["departamento"]] = reemplazos["{{DEPARTAMENTO}}"]
            if mapeo.get("edificio"): ws[mapeo["edificio"]] = reemplazos["{{EDIFICIO}}"]
            if mapeo.get("periodo"): ws[mapeo["periodo"]] = reemplazos["{{PERIODO}}"]
            if mapeo.get("fecha"): ws[mapeo["fecha"]] = reemplazos["{{FECHA}}"]
            if mapeo.get("total"): ws[mapeo["total"]] = reemplazos["{{TOTAL}}"]
            
            fila_inicio = mapeo.get("tabla_fila")
            if fila_inicio and fila_inicio.isdigit():
                fila_inicio_tabla = int(fila_inicio)
                for i, item in enumerate(items):
                    r = fila_inicio_tabla + i
                    if mapeo.get("tabla_col_cant"): ws[f"{mapeo['tabla_col_cant']}{r}"] = item['cantidad']
                    if mapeo.get("tabla_col_cat"): ws[f"{mapeo['tabla_col_cat']}{r}"] = item['categoria']
                    if mapeo.get("tabla_col_concepto"): ws[f"{mapeo['tabla_col_concepto']}{r}"] = item['concepto']
                    if mapeo.get("tabla_col_precio"): ws[f"{mapeo['tabla_col_precio']}{r}"] = item['precio_unitario']
                    if mapeo.get("tabla_col_subtotal"): ws[f"{mapeo['tabla_col_subtotal']}{r}"] = item['subtotal']
        else:
            cols_tabla = {}
            fila_inicio_tabla = -1
            
            # Escanear hoja buscando palabras clave (Fallback)
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        val = cell.value
                        # Variables simples
                        for clave, valor in reemplazos.items():
                            if clave in val:
                                val = val.replace(clave, str(valor))
                                cell.value = val
                        
                        # Detectar columnas de la tabla
                        if "{{T_CANT}}" in str(cell.value):
                            cols_tabla['cant'] = cell.column
                            if fila_inicio_tabla == -1: fila_inicio_tabla = cell.row
                            cell.value = str(cell.value).replace("{{T_CANT}}", "")
                        if "{{T_CAT}}" in str(cell.value):
                            cols_tabla['cat'] = cell.column
                            if fila_inicio_tabla == -1: fila_inicio_tabla = cell.row
                            cell.value = str(cell.value).replace("{{T_CAT}}", "")
                        if "{{T_CONCEPTO}}" in str(cell.value):
                            cols_tabla['concepto'] = cell.column
                            if fila_inicio_tabla == -1: fila_inicio_tabla = cell.row
                            cell.value = str(cell.value).replace("{{T_CONCEPTO}}", "")
                        if "{{T_PRECIO}}" in str(cell.value):
                            cols_tabla['precio'] = cell.column
                            if fila_inicio_tabla == -1: fila_inicio_tabla = cell.row
                            cell.value = str(cell.value).replace("{{T_PRECIO}}", "")
                        if "{{T_SUBTOTAL}}" in str(cell.value):
                            cols_tabla['subtotal'] = cell.column
                            if fila_inicio_tabla == -1: fila_inicio_tabla = cell.row
                            cell.value = str(cell.value).replace("{{T_SUBTOTAL}}", "")
                            
            # Escribir los items respetando el formato visual de las celdas
            if fila_inicio_tabla != -1:
                for i, item in enumerate(items):
                    r = fila_inicio_tabla + i
                    if 'cant' in cols_tabla: ws.cell(row=r, column=cols_tabla['cant'], value=item['cantidad'])
                    if 'cat' in cols_tabla: ws.cell(row=r, column=cols_tabla['cat'], value=item['categoria'])
                    if 'concepto' in cols_tabla: ws.cell(row=r, column=cols_tabla['concepto'], value=item['concepto'])
                    if 'precio' in cols_tabla: ws.cell(row=r, column=cols_tabla['precio'], value=item['precio_unitario'])
                    if 'subtotal' in cols_tabla: ws.cell(row=r, column=cols_tabla['subtotal'], value=item['subtotal'])
                
    # --- FLUJO: SIN PLANTILLA (DEFAULT) ---
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recibo"
        
        ws.merge_cells('A1:E2')
        ws['A1'] = f"RECIBO - {recibo['edificio_nombre']}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A4'] = "Departamento:"
        ws['B4'] = recibo['identificador']
        ws['A5'] = "Inquilino:"
        ws['B5'] = recibo['inquilino']
        ws['D4'] = "Período:"
        ws['E4'] = recibo['periodo']
        
        headers = ["Cant", "Categoría", "Concepto", "Precio", "Subtotal"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=8, column=col, value=h)
            c.font = Font(bold=True)
            
        fila = 9
        for item in items:
            ws.cell(row=fila, column=1, value=item['cantidad'])
            ws.cell(row=fila, column=2, value=item['categoria'])
            ws.cell(row=fila, column=3, value=item['concepto'])
            ws.cell(row=fila, column=4, value=item['precio_unitario'])
            ws.cell(row=fila, column=5, value=item['subtotal'])
            fila += 1
            
        ws.cell(row=fila+1, column=4, value="TOTAL:")
        ws.cell(row=fila+1, column=5, value=recibo['total']).font = Font(bold=True)

    # Guardar
    try:
        wb.save(ruta_archivo)
    except PermissionError:
        ruta_archivo = ruta_base / f"{nombre_archivo.replace('.xlsx', '')}_{int(time.time())}.xlsx"
        wb.save(ruta_archivo)

    # Conversión PDF
    formato_salida = config.get("formato_salida", "Solo Excel (.xlsx)")
    ruta_final = str(ruta_archivo)
    if formato_salida == "Excel y PDF":
        try:
            import shutil
            comando_base = "libreoffice"
            if platform.system() == "Windows":
                comando_base = "soffice"
                if not shutil.which(comando_base):
                    rutas_comunes = [
                        r"C:\Program Files\LibreOffice\program\soffice.exe",
                        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"
                    ]
                    for ruta in rutas_comunes:
                        if os.path.exists(ruta):
                            comando_base = ruta
                            break
            
            comando = [comando_base, "--headless", "--nologo", "--nofirststartwizard", "--convert-to", "pdf", str(ruta_archivo), "--outdir", str(ruta_base)]
            subprocess.run(comando, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            ruta_final = str(ruta_archivo).replace('.xlsx', '.pdf')
        except Exception as e:
            print(f"Error generando PDF: {e}")

    database.actualizar_recibo(recibo_id, archivo_excel=ruta_final, total=recibo['total'])
    return ruta_final

def abrir_archivo(ruta_archivo):
    if not os.path.exists(ruta_archivo): return False
    try:
        if platform.system() == 'Darwin': subprocess.call(('open', ruta_archivo))
        elif platform.system() == 'Windows': os.startfile(ruta_archivo)
        else: subprocess.call(('xdg-open', ruta_archivo))
        return True
    except: return False
